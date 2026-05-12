import sqlite3
import csv
import os
import openpyxl
from pathlib import Path


def clean_text(value):
    if value is None:
        return ""
    text = str(value).replace('ـ', '')
    return text.strip()


def create_sqlite_from_csv(csv_file_path, db_name, table_name):
    # 1. Connect to database (will be created if it doesn't exist)
    conn = sqlite3.connect(db_name)
    cursor = conn.cursor()

    try:
        # 2. Open CSV file with correct encoding for Arabic and semicolon delimiter
        with open(csv_file_path, mode='r', encoding='utf-8-sig') as f:
            # Specify delimiter=';' because Arabic Excel uses it
            reader = csv.reader(f, delimiter=';')
            
            # Read first row (headers)
            headers = next(reader)
            
            # 3. Clean column names (remove spaces and symbols to be SQL compatible)
            clean_headers = []
            for h in headers:
                clean_h = h.strip().replace('/', '_').replace(' ', '_').replace(';', '_')
                if not clean_h: clean_h = "column_unknown"
                clean_headers.append(clean_h)

            # 4. Create table based on cleaned names
            # Assume all columns are TEXT for simplicity
            cols_query = ", ".join([f'"{name}" TEXT' for name in clean_headers])
            cursor.execute(f"DROP TABLE IF EXISTS {table_name}")
            cursor.execute(f"CREATE TABLE {table_name} ({cols_query})")

            # 5. Insert data
            insert_query = f"INSERT INTO {table_name} VALUES ({', '.join(['?' for _ in clean_headers])})"
            
            count = 0
            for row in reader:
                if any(row):  # Ensure row is not empty
                    cursor.execute(insert_query, row)
                    count += 1

            conn.commit()
            print(f"✅ Success! Database created: {db_name}")
            print(f"✅ Imported {count} records to table: {table_name}")
            print(f"✅ New column names: {', '.join(clean_headers)}")

    except Exception as e:
        print(f"❌ Error: {e}")
    finally:
        conn.close()

def create_sqlite_from_xlsx(xlsx_file_path, db_name):
    """تحويل ملف XLSX إلى قاعدة بيانات SQLite مع دعم جداول متعددة"""
    try:
        workbook = openpyxl.load_workbook(xlsx_file_path)
        conn = sqlite3.connect(db_name)
        cursor = conn.cursor()
        
        print(f"\n📄 معالجة الملف: {xlsx_file_path}")
        print(f"   أوراق العمل المتوفرة: {workbook.sheetnames}")
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # استخراج البيانات والعناوين
            rows = list(worksheet.iter_rows(values_only=True))
            if not rows:
                continue
            
            headers = [clean_text(h) if h else "column_unknown" for h in rows[0]]
            
            # تنظيف أسماء الأعمدة
            clean_headers = []
            for h in headers:
                clean_h = h.replace('/', '_').replace(' ', '_').replace(';', '_')
                if not clean_h:
                    clean_h = "column_unknown"
                clean_headers.append(clean_h)
            
            # إنشاء اسم الجدول من اسم ورقة العمل
            table_name = "company_phones"  # استخدام اسم الجدول بالإنجليزية
            
            # إنشاء الجدول بالهيكل المحدد
            cursor.execute(f"DROP TABLE IF EXISTS {table_name}")
            cursor.execute(f"""
                CREATE TABLE {table_name} (
                    office_company      TEXT,
                    house_company       TEXT,
                    offices_civil       TEXT,
                    house_civil         TEXT,
                    division_name       TEXT,
                    address             TEXT,
                    authority           TEXT,
                    department          TEXT,
                    note                TEXT
                )
            """)
            
            # إدراج البيانات مع التنظيف والفصل
            count = 0
            for row in rows[1:]:
                if any(row):  # تأكد من أن الصف ليس فارغاً
                    # تحويل جميع القيم إلى نصوص وإزالة tatweel
                    row_data = [clean_text(cell) for cell in row]
                    
                    # تخطي العمود الأول (number) واستخدام الأعمدة من 2 إلى 8
                    row_data = row_data[1:8]  # تخطي العمود الأول
                    
                    # إذا كان عدد الأعمدة أقل من 7، أضف أعمدة فارغة
                    while len(row_data) < 7:
                        row_data.append("")
                    
                    # اقتصاص على 7 أعمدة فقط
                    row_data = row_data[:7]
                    
                    # استخراج جميع الأعمدة
                    office_company = row_data[0]
                    house_company = row_data[1]
                    offices_civil = row_data[2]
                    house_civil = row_data[3]
                    name_address = row_data[4]
                    department_division = row_data[5]
                    note = row_data[6]
                    
                    # فصل name_address على "/"
                    if '/' in name_address:
                        division_name, address = name_address.split('/', 1)
                        division_name = division_name.strip()
                        address = address.strip()
                    else:
                        division_name = name_address
                        address = ""
                    
                    # فصل department_division على "/"
                    if '/' in department_division:
                        authority, department = department_division.split('/', 1)
                        authority = authority.strip()
                        department = department.strip()
                    else:
                        authority = department_division
                        department = ""
                    
                    cursor.execute(f"""
                        INSERT INTO {table_name} 
                        (office_company, house_company, offices_civil, house_civil, division_name, address, authority, department, note) 
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (office_company, house_company, offices_civil, house_civil, division_name, address, authority, department, note))
                    count += 1
            
            conn.commit()
            print(f"   ✅ تم إنشاء جدول: {table_name} ({count} سجل بعد التنظيف)")
        
        conn.close()
        print(f"✅ تم إنشاء قاعدة البيانات: {db_name}")
        
    except Exception as e:
        print(f"❌ خطأ في معالجة {xlsx_file_path}: {e}")

def convert_all_xlsx_to_sqlite(directory="."):
    """تحويل جميع ملفات XLSX في المجلد إلى قواعد بيانات SQLite"""
    xlsx_files = list(Path(directory).glob("*.xlsx"))
    
    if not xlsx_files:
        print(f"⚠️ لم يتم العثور على ملفات XLSX في المجلد: {directory}")
        return
    
    print(f"🔍 تم العثور على {len(xlsx_files)} ملف XLSX")
    
    for xlsx_file in xlsx_files:
        # استخدام اسم قاعدة البيانات بالإنجليزية
        db_name = "company_phones.db"
        db_path = xlsx_file.parent / db_name
        
        create_sqlite_from_xlsx(str(xlsx_file), str(db_path))

# --- تشغيل السكريبت ---
if __name__ == "__main__":
    # تحويل جميع ملفات XLSX في المجلد الحالي
    convert_all_xlsx_to_sqlite(".")